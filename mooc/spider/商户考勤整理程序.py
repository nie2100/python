# -*- coding: utf-8 -*-
# 此程序作者：聂光明
from openpyxl import load_workbook, Workbook

# 打开excel文件
data = load_workbook('未处理.xlsx')
# 创建新的excel
newdata = Workbook()
# 获取excel的sheet
dsn = data.sheetnames
v = []
#  在新excel创建新sheet
for sheet in dsn:
    newdata.create_sheet(sheet)
    for row in data[sheet].iter_rows():
        print(v)
        if len(v):
            print(row[0].value)
            if row[0].value == v[0]:
                v1 = []
                for i in row:
                    if i.value:
                        v1.append(i.value)
                v = v + v1[1:]
            else:
                newdata[sheet].append(v)
                v = []
                for i in row:
                    # print(i.value)
                    if i.value:
                        v.append(i.value)
        else:
            newdata[sheet].append(v)
            for i in row:
                # print(i.value)
                if i.value:
                    v.append(i.value)

# 保存新的excel
newdata.save('整理完毕.xlsx')
